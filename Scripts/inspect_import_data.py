from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import hashlib

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1] / "Import Data"


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def non_empty(values: tuple[object, ...]) -> list[str]:
    return ["" if value is None else str(value) for value in values if value is not None and str(value).strip() != ""]


def workbook_preview(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    print(f"FILE: {path.name}")
    print(f"  sheets: {', '.join(workbook.sheetnames)}")
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_row = None
        header_count = -1
        for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(20, worksheet.max_row), values_only=True), start=1):
            values = non_empty(row)
            if len(values) > header_count:
                header_count = len(values)
                header_row = (row_index, values)
        if header_row:
            row_index, values = header_row
            print(f"  - {sheet_name}: rows={worksheet.max_row}, cols={worksheet.max_column}, header_row={row_index}, header={values[:22]}")
        else:
            print(f"  - {sheet_name}: rows={worksheet.max_row}, cols={worksheet.max_column}, header_row=none")
    print()


def main() -> None:
    xlsx_files = sorted([path for path in ROOT.iterdir() if path.is_file() and path.suffix.lower() == ".xlsx"])
    print(f"XLSX files: {len(xlsx_files)}")
    print()

    hash_groups: dict[str, list[str]] = defaultdict(list)
    for path in xlsx_files:
        hash_groups[file_hash(path)].append(path.name)

    print("Duplicate groups by file hash:")
    duplicate_groups = [names for names in hash_groups.values() if len(names) > 1]
    if duplicate_groups:
        for names in duplicate_groups:
            print(" - " + " | ".join(names))
    else:
        print(" - none")
    print()

    for path in xlsx_files:
        workbook_preview(path)


if __name__ == "__main__":
    main()
