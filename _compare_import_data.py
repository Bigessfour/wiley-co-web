from pathlib import Path
import csv
import hashlib
import zipfile
from openpyxl import load_workbook

base = Path('c:/Users/biges/Desktop/WW AWS/Import Data')
files = [p for p in sorted(base.iterdir()) if p.is_file() and p.name.lower() != 'wsddata.zip']


def normalize_cell(value):
    if value is None:
        return ''
    return str(value).strip()


def csv_signature(path):
    with path.open('r', encoding='utf-8-sig', newline='') as handle:
        rows = [tuple(normalize_cell(cell) for cell in row) for row in csv.reader(handle)]
    digest = hashlib.sha256(repr(rows).encode('utf-8')).hexdigest()
    return digest, len(rows), len(rows[0]) if rows else 0, rows[:3]


def xlsx_signature(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    preview = []
    for worksheet in workbook.worksheets:
        values = []
        for row in worksheet.iter_rows(values_only=True):
            values.append(tuple(normalize_cell(cell) for cell in row))
        sheets.append((worksheet.title, values))
        preview.append((worksheet.title, values[:3]))
    digest = hashlib.sha256(repr(sheets).encode('utf-8')).hexdigest()
    return digest, len(workbook.worksheets), preview

summary = []
for path in files:
    if path.suffix.lower() == '.csv':
        signature, row_count, column_count, preview = csv_signature(path)
        summary.append((path.name, 'csv', signature, row_count, column_count, preview))
    elif path.suffix.lower() == '.xlsx':
        signature, sheet_count, preview = xlsx_signature(path)
        summary.append((path.name, 'xlsx', signature, sheet_count, None, preview))
    else:
        summary.append((path.name, 'other', None, None, None, None))

print('FILE_SUMMARY')
for item in summary:
    print(item[0])
    print('  type:', item[1])
    print('  sig :', item[2])
    if item[1] == 'csv':
        print('  rows:', item[3], 'cols:', item[4])
    else:
        print('  sheets:', item[3])
    print('  preview:', item[5])

print('\nDUPLICATE_GROUPS')
groups = {}
for item in summary:
    groups.setdefault((item[1], item[2]), []).append(item[0])
for key, names in groups.items():
    if len(names) > 1:
        print(key, names)

zip_path = base / 'wsddata.zip'
print('\nZIP_CONTENTS')
with zipfile.ZipFile(zip_path) as zf:
    for info in zf.infolist():
        print(info.filename, info.file_size)
